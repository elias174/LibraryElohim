import sys
import openpyxl

from models import *

# Base = declarative_base()
#
# db = create_engine('sqlite:///dataBase.db', echo=False, encoding='utf8')
# metadata = MetaData(db)
#
# Session = sessionmaker(bind=db)
#
# session = Session()

reload(sys)
sys.setdefaultencoding('utf8')


class NotFilledError(Exception):
    pass

class ExportXLSXError(Exception):
    pass


class Adapter_XLSX:
    def __init__(self, file_name):
        self.default_categoria = self.get_or_create_category('Utiles Libreria')
        self.productos = []
        self.dict_data = {}
        self.generate_sheet(file_name)
        self.modified = False

    def get_or_create_category(self, name):
        instance = session.query(Categoria).filter_by(nombre=name).first()
        if instance:
            return instance
        else:
            category = Categoria(
                name, 'Venta libreria')
            session.add(category)
            session.flush()
            session.refresh(category)
            return category

    def generate_sheet(self, name_file):
        self.wb = openpyxl.load_workbook(name_file)
        self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])

    def extract_data(self):
        # assert(self.categoria)
        for row in range(2, self.sheet.max_row + 1):
            id = self.sheet['A' + str(row)].value
            articulo = self.sheet['B' + str(row)].value
            cantidad = self.sheet['C' + str(row)].value
            p_compra = self.sheet['D' + str(row)].value
            p_venta = self.sheet['E' + str(row)].value
            categoria = self.sheet['F' + str(row)].value

            if id and isinstance(id, long):
                product = session.query(Producto).get(int(id))
                if not product:
                    raise NotFilledError(
                        'Revisar fila %s columna ID no valido' % str(row))
                if cantidad < 0:
                    raise NotFilledError(
                        'Revisar fila %s columna Cantidad' % str(row))
                product.stock += int(cantidad)
                self.modified = True
                continue

            if not articulo:
                raise NotFilledError(
                    'Revisar fila %s columna Articulo' % str(row))
            if cantidad < 0:
                raise NotFilledError(
                    'Revisar fila %s columna Cantidad' % str(row))
            if p_compra < float(0):
                raise NotFilledError(
                    'Revisar fila %s columna P.Compra' % str(row))
            if p_venta < float(0):
                raise NotFilledError(
                    'Revisar fila %s columna P.Venta' % str(row))

            articulo = articulo.capitalize()
            categoria_id = self.default_categoria.id
            if categoria:
                categoria_id = self.get_or_create_category(categoria).id
            producto = Producto(
                categoria_id,
                articulo,
                float(p_compra),
                float(p_venta),
                int(cantidad),
                'Producto Nuevo'
            )
            self.productos.append(producto)

    def save_products(self):
        self.extract_data()
        if len(self.productos) <= 0 and not self.modified:
            raise NotFilledError('No existe ningun Producto en su XLSX')
        for producto in self.productos:
            session.add(producto)
        session.commit()

    @staticmethod
    def export_products_xlsx(filename):
        products = (session.query(Producto).order_by(Producto.categoria).all())
        if len(products) < 1:
            raise ExportXLSXError('No tiene inventario registrado')
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'InventarioLibreria'
            sheet['A1'].value = 'ID'
            sheet['B1'].value = 'Nombre'
            sheet['C1'].value = 'Cantidad'
            sheet['D1'].value = 'P.Compra'
            sheet['E1'].value = 'P.Venta'
            sheet['F1'].value = 'Detalle'
            sheet['G1'].value = 'Categoria'

            index = 2
            for product in products:
                sheet['A' + str(index)].value = product.id
                sheet['B' + str(index)].value = str(product.nombre).decode('iso-8859-1')
                sheet['C' + str(index)].value = product.stock
                sheet['D' + str(index)].value = float(product.precio_compra)
                sheet['E' + str(index)].value = float(product.precio_venta)
                sheet['F' + str(index)].value = str(product.detalle).decode('iso-8859-1')
                sheet['G' + str(index)].value = (
                    session.query(Categoria).get(product.categoria).nombre)
                index += 1

            wb.save(filename=filename)
        except:
            raise ExportXLSXError('Algo Salio Mal, lo sentimos')
